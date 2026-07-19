"""Structured load: xlsx registers -> typed graph nodes. NO LLM involved (spec 8 P1).

Run:  py pipeline/load_structured.py     (after schema.py)

Handles the injected messiness deliberately:
  - tag spellings resolve through resolve.canonical_key; non-register spellings
    (e.g. "P101-A") are recorded on Equipment.aliases
  - dates parse from both YYYY-MM-DD and DD/MM/YYYY
  - unknown failure codes (the planted "LKE" typo) are corrected to the closest
    vocabulary code (anagram or edit-distance 1) and the raw value is preserved
"""

import sys
from datetime import datetime
from pathlib import Path

from neo4j import GraphDatabase
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from config import NEO4J_URI, NEO4J_USER, NEO4J_PASSWORD, NEO4J_DATABASE, CORPUS
from resolve import canonical_key

FAILURE_MODES = {
    "SER": "Seal leakage", "VIB": "Excessive vibration", "BRG": "Bearing failure",
    "PLU": "Plugged / choked", "LEK": "External leakage", "OHE": "Overheating",
    "ELP": "Electrical problem",
}

SITE = "Plant Site A"
AREAS = {"Area-1": "Process Area", "Area-2": "Utilities Area"}


def rows(path):
    ws = load_workbook(path, read_only=True).active
    it = ws.iter_rows(values_only=True)
    header = [str(h) for h in next(it)]
    for r in it:
        if any(v is not None for v in r):
            yield dict(zip(header, r))


def parse_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"unparseable date: {v!r}")


def fix_code(raw):
    """Return (canonical_code, was_corrected). Closed 7-code vocabulary."""
    if raw in FAILURE_MODES:
        return raw, False
    for code in FAILURE_MODES:
        if sorted(code) == sorted(raw):          # transposition typo (LKE -> LEK)
            return code, True
        diffs = sum(a != b for a, b in zip(code, raw)) + abs(len(code) - len(raw))
        if diffs <= 1:                            # single-char typo
            return code, True
    raise ValueError(f"cannot map failure code {raw!r}")


def main():
    equipment = list(rows(CORPUS / "equipment_register.xlsx"))
    tag_by_key = {canonical_key(e["tag"]): e["tag"] for e in equipment}

    def resolve_tag(raw):
        """raw spelling -> (register_tag, alias_or_None)."""
        key = canonical_key(str(raw))
        tag = tag_by_key.get(key) if key else None
        if tag is None:
            raise ValueError(f"unresolvable equipment tag: {raw!r}")
        return tag, (str(raw) if str(raw) != tag else None)

    wos, corrected = [], 0
    for r in rows(CORPUS / "work_orders.xlsx"):
        tag, alias = resolve_tag(r["equipment_tag"])
        code, raw_code = None, None
        if r["failure_code"]:
            code, was = fix_code(str(r["failure_code"]).strip())
            raw_code = str(r["failure_code"]).strip() if was else None
            corrected += was
        wos.append({
            "wo_id": r["wo_id"], "tag": tag, "alias": alias,
            "date": parse_date(r["date"]), "type": r["type"], "code": code,
            "raw_code": raw_code, "downtime_hrs": float(r["downtime_hrs"] or 0),
            "description": r["description"], "technician": r["technician"],
        })

    inspections = []
    for r in rows(CORPUS / "inspection_records.xlsx"):
        tag, alias = resolve_tag(r["equipment_tag"])
        inspections.append({
            "record_id": r["record_id"], "tag": tag, "alias": alias,
            "date": parse_date(r["inspection_date"]), "result": r["result"],
            "valid_until": parse_date(r["valid_until"]),
        })

    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PASSWORD))
    with driver.session(database=NEO4J_DATABASE) as s:
        s.run("MERGE (:Site {name: $n})", n=SITE)
        for area, desc in AREAS.items():
            s.run("MATCH (st:Site {name: $s}) MERGE (a:Area {name: $a}) "
                  "SET a.description = $d MERGE (st)-[:CONTAINS]->(a)",
                  s=SITE, a=area, d=desc)
        s.run("""
            UNWIND $rows AS r
            MATCH (a:Area {name: r.area})
            MERGE (e:Equipment {tag: r.tag})
            SET e.iso14224_class = r.iso14224_class, e.criticality = r.criticality,
                e.service = r.service,
                e.aliases = coalesce(e.aliases, [r.tag])
            MERGE (a)-[:CONTAINS]->(e)
        """, rows=equipment)
        s.run("""
            UNWIND $modes AS m
            MERGE (f:FailureMode {code: m.code}) SET f.description = m.description
        """, modes=[{"code": c, "description": d} for c, d in FAILURE_MODES.items()])
        s.run("""
            UNWIND $rows AS r
            MATCH (e:Equipment {tag: r.tag})
            MERGE (w:WorkOrder {wo_id: r.wo_id})
            SET w.date = date(r.date), w.type = r.type,
                w.downtime_hrs = r.downtime_hrs, w.description = r.description,
                w.technician = r.technician, w.raw_failure_code = r.raw_code
            MERGE (w)-[:PERFORMED_ON]->(e)
            FOREACH (_ IN CASE WHEN r.alias IS NULL THEN [] ELSE [1] END |
                SET e.aliases = CASE WHEN r.alias IN e.aliases
                                     THEN e.aliases ELSE e.aliases + r.alias END)
            FOREACH (_ IN CASE WHEN r.code IS NULL THEN [] ELSE [1] END |
                MERGE (f:FailureMode {code: r.code})
                MERGE (w)-[:RESULTED_IN]->(f))
        """, rows=wos)
        s.run("""
            UNWIND $rows AS r
            MATCH (e:Equipment {tag: r.tag})
            MERGE (i:InspectionRecord {record_id: r.record_id})
            SET i.date = date(r.date), i.result = r.result,
                i.valid_until = date(r.valid_until)
            MERGE (i)-[:ON]->(e)
            FOREACH (_ IN CASE WHEN r.alias IS NULL THEN [] ELSE [1] END |
                SET e.aliases = CASE WHEN r.alias IN e.aliases
                                     THEN e.aliases ELSE e.aliases + r.alias END)
        """, rows=inspections)

        counts = s.run("""
            MATCH (n) WITH labels(n)[0] AS label, count(*) AS n
            RETURN label, n ORDER BY label
        """).data()
        rels = s.run("MATCH ()-[r]->() RETURN type(r) AS t, count(*) AS n ORDER BY t").data()

    driver.close()
    print(f"loaded {len(equipment)} equipment, {len(wos)} work orders "
          f"({corrected} failure code corrected), {len(inspections)} inspections")
    print("nodes:", {c["label"]: c["n"] for c in counts})
    print("rels :", {r["t"]: r["n"] for r in rels})


if __name__ == "__main__":
    main()
